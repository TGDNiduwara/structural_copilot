"""
tools/excel_tool.py
====================
Excel report/workbook generation engine for structural analysis results.

Builds a professional, multi-sheet .xlsx workbook (Member Forces, Reactions,
Bill of Quantities) using pandas for data handling and openpyxl for styling,
formulas, and layout polish.

Requires: pandas, openpyxl (`pip install pandas openpyxl`)
"""

from __future__ import annotations

import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("structural_copilot.excel_tool")
logger.setLevel(logging.INFO)


# --------------------------------------------------------------------------
# Shared style constants
# --------------------------------------------------------------------------

BANNER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
BANNER_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="D9D9D9")

HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_SIDE = Side(style="thin", color="BFBFBF")
GRID_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")

SUMMARY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUMMARY_FONT = Font(name="Calibri", size=11, bold=True)

NUMERIC_COLUMNS_2DP = {
    "Position_m",
    "FX_kN",
    "FZ_kN",
    "MY_kNm",
    "Total_Length_m",
    "Unit_Mass_kg_m",
    "Total_Weight_kg",
}


class ExcelReporter:
    """Generates professionally formatted structural-engineering workbooks."""

    def __init__(self):
        self.workbook: Workbook | None = None
        # [FIX M13] Track used sheet names to avoid duplicates after truncation
        self._used_sheet_names: set[str] = set()

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def create_structural_workbook(
        self,
        file_path: str,
        project_name: str,
        member_forces_df: pd.DataFrame,
        reactions_df: pd.DataFrame,
        boq_df: pd.DataFrame,
    ) -> str:
        """
        Builds and saves a 3-sheet workbook: 'Member Forces', 'Reactions',
        and 'Bill of Quantities'. Returns the saved file path.
        """
        self.workbook = Workbook()
        self._used_sheet_names = set()
        # Remove the default blank sheet; we add our own in order.
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        self._build_sheet(
            title="Member Forces",
            banner_title=f"{project_name} — Member Force Envelope",
            df=member_forces_df,
            summary_rules=[("MY_kNm", "max", "Governing Moment (kNm)")],
        )

        self._build_sheet(
            title="Reactions",
            banner_title=f"{project_name} — Support Reactions",
            df=reactions_df,
            summary_rules=[
                ("FZ_kN", "sum", "Total Vertical Reaction (kN)"),
            ],
        )

        self._build_sheet(
            title="Bill of Quantities",
            banner_title=f"{project_name} — Bill of Quantities (Steel)",
            df=boq_df,
            summary_rules=[("Total_Weight_kg", "sum", "Total Structural Steel Weight (kg)")],
            use_excel_sum_formula=True,
        )

        self.workbook.save(file_path)
        logger.info("Workbook saved to %s", file_path)
        return file_path

    def build_workbook_from_sheets(
        self,
        file_path: str,
        sheets: dict[str, pd.DataFrame],
        use_excel_sum_formula: bool = False,
    ) -> str:
        """
        [WP6] Builds a workbook from a dict of {sheet_title: DataFrame} —
        the LLM chooses which sheets to include (member forces, reactions,
        displacements, stresses, BOQ, modal, ...). Reuses the standard
        styling. Returns the saved file path.
        """
        self.workbook = Workbook()
        self._used_sheet_names = set()
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        for title, df in sheets.items():
            self._build_sheet(
                title=title[:28],
                banner_title=title.replace("_", " "),
                df=df,
                summary_rules=[],
                use_excel_sum_formula=use_excel_sum_formula,
            )
        self.workbook.save(file_path)
        logger.info("Workbook (custom sheets) saved to %s", file_path)
        return file_path

    # ------------------------------------------------------------------ #
    # Internal sheet builder
    # ------------------------------------------------------------------ #

    def _deduplicate_sheet_name(self, title: str) -> str:
        """[FIX M13] Ensures sheet names are unique after 31-char truncation."""
        name = title[:31]  # Excel sheet name limit
        if name not in self._used_sheet_names:
            self._used_sheet_names.add(name)
            return name
        # Append a counter suffix
        counter = 2
        while True:
            suffix = f" ({counter})"
            candidate = title[: 31 - len(suffix)] + suffix
            if candidate not in self._used_sheet_names:
                self._used_sheet_names.add(candidate)
                return candidate
            counter += 1

    def _build_sheet(
        self,
        title: str,
        banner_title: str,
        df: pd.DataFrame,
        summary_rules: list | None = None,
        use_excel_sum_formula: bool = False,
    ) -> Worksheet:
        sheet_name = self._deduplicate_sheet_name(title)
        ws = self.workbook.create_sheet(title=sheet_name)
        n_cols = max(len(df.columns), 1)

        # ---- Banner header (project title) ----
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        banner_cell = ws.cell(row=1, column=1, value=banner_title)
        banner_cell.font = BANNER_FONT
        banner_cell.fill = BANNER_FILL
        banner_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        subtitle_cell = ws.cell(
            row=2,
            column=1,
            value=f"Generated by Structural Multi-App Agent — "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )
        subtitle_cell.font = Font(name="Calibri", size=9, italic=True, color="595959")
        subtitle_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16

        header_row = 4

        # ---- Column headers ----
        if df.empty:
            ws.cell(row=header_row, column=1, value="No data available for this section.")
            ws.cell(row=header_row, column=1).font = Font(italic=True, color="808080")
            return ws

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=str(col_name).replace("_", " "))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = GRID_BORDER

        # ---- Data rows ----
        first_data_row = header_row + 1
        for row_offset, (_, row) in enumerate(df.iterrows()):
            excel_row = first_data_row + row_offset
            for col_idx, col_name in enumerate(df.columns, start=1):
                value = row[col_name]
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGN
                cell.border = GRID_BORDER
                if col_name in NUMERIC_COLUMNS_2DP and isinstance(value, (int, float)):
                    cell.number_format = "0.00"

        last_data_row = first_data_row + len(df) - 1

        # ---- Auto-adjust column widths ----
        for col_idx, col_name in enumerate(df.columns, start=1):
            # Limit the number of rows sampled for width calculation
            sample_values = df[col_name].head(100).tolist()
            max_len = max(
                [len(str(col_name))]
                + [len(f"{v:.2f}" if isinstance(v, float) else str(v)) for v in sample_values]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 3, 12)

        # ---- Summary formulas block ----
        if summary_rules:
            summary_row = last_data_row + 2
            for label_col_name, agg, label in summary_rules:
                if label_col_name not in df.columns:
                    continue
                col_idx = list(df.columns).index(label_col_name) + 1
                col_letter = get_column_letter(col_idx)

                label_cell = ws.cell(row=summary_row, column=1, value=label)
                label_cell.font = SUMMARY_FONT
                label_cell.fill = SUMMARY_FILL

                value_cell = ws.cell(row=summary_row, column=col_idx)
                if use_excel_sum_formula and agg == "sum":
                    value_cell.value = (
                        f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                    )
                elif agg == "sum":
                    value_cell.value = float(df[label_col_name].sum())
                elif agg == "max":
                    value_cell.value = (
                        f"=MAX({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                    )
                elif agg == "min":
                    value_cell.value = (
                        f"=MIN({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                    )

                value_cell.font = SUMMARY_FONT
                value_cell.fill = SUMMARY_FILL
                value_cell.number_format = "0.00"
                value_cell.border = GRID_BORDER
                summary_row += 1

        ws.freeze_panes = ws.cell(row=first_data_row, column=1)
        return ws
